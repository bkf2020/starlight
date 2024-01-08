import django
from django.conf import settings
import os
import sys

sys.path.insert(0, '../../')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'starlight.settings')
django.setup()

from problems.models import Problem
from django.contrib.postgres.search import TrigramSimilarity, TrigramWordSimilarity, TrigramStrictWordSimilarity, TrigramDistance
import openpyxl
from urllib.parse import urlparse
from urllib.parse import parse_qs

full_spreadsheet = openpyxl.load_workbook('Competitive Programming Progress.xlsx')

"""
sheet_names = full_spreadsheet.sheetnames
for name in sheet_names:
"""

for curr_spreadsheet in full_spreadsheet:
    for i in range(3, curr_spreadsheet.max_row):
        # https://docs.djangoproject.com/en/4.2/ref/contrib/postgres/search/
        current_cell = curr_spreadsheet.cell(row=i, column=1)
        if(current_cell.value is None or current_cell.value == ""):
            continue
        
        if current_cell.hyperlink is not None:
            desired_similarity=0.9
            parsed = urlparse(current_cell.hyperlink.target)
            try:
                problem_id = parse_qs(parsed.query)['cpid'][0]
            except:
                problem_id = 0
            if "usaco" in current_cell.hyperlink.target.lower() and int(problem_id) < 487:
                continue
            if "codeforces" in current_cell.hyperlink.target.lower():
                desired_similarity=0.7
            desired_problems = Problem.objects.annotate(
                similarity=TrigramSimilarity("link", current_cell.hyperlink.target),
            ).filter(
                similarity__gte=desired_similarity
            ).order_by("-similarity")
        else:
            desired_problems = Problem.objects.annotate(
                similarity=TrigramWordSimilarity(current_cell.value, "name"),
            ).filter(
                similarity__gt=0.6
            ).order_by("-similarity")
        if desired_problems.count() > 0:
            problem_name = desired_problems.first().name
            insights_str = curr_spreadsheet.cell(row=i, column=3).value
            insights_arr = insights_str.split("*")
            insights = []
            for insight in insights_arr:
                if insight.strip() != "":
                    insights.append(insight.strip())
            print(insights)
            print(desired_problems.first().name)